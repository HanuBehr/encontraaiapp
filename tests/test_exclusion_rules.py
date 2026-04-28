from __future__ import annotations

from io import BytesIO
from pathlib import Path
import sys

from openpyxl import load_workbook

from app.enums import LeadSourceType, LeadStatus
from app.models.lead import Lead
from app.models.organization import Organization
from app.repositories.lead_repository import LeadRepository
from app.schemas.discovery import DiscoveryLeadCandidate
from app.schemas.lead import LeadListFilters
from app.services.exclusion_rules import ExclusionRuleService
from app.services.export_excel import ExcelExportService
from app.services.normalization import normalize_business_name
from scripts import import_exclusion_rules


def _organization(slug: str = "tenant-a") -> Organization:
    return Organization(slug=slug, name=slug.title(), display_name=slug.title())


def _lead(
    business_name: str,
    *,
    organization: Organization,
    domain: str | None = None,
    category: str | None = "loja de tintas",
    is_blocked: bool = False,
    blocked_reason: str | None = None,
) -> Lead:
    return Lead(
        organization=organization,
        business_name=business_name,
        normalized_business_name=normalize_business_name(business_name) or business_name.lower(),
        category=category,
        city="Campinas",
        state="SP",
        website=f"https://{domain}" if domain else None,
        domain=domain,
        lead_source_type=LeadSourceType.GOOGLE_PLACES,
        status=LeadStatus.NEW,
        is_blocked=is_blocked,
        blocked_reason=blocked_reason,
    )


def _candidate(
    business_name: str,
    *,
    domain: str | None = None,
    category: str | None = "loja de tintas",
) -> DiscoveryLeadCandidate:
    return DiscoveryLeadCandidate(
        business_name=business_name,
        normalized_business_name=normalize_business_name(business_name) or business_name.lower(),
        category=category,
        city="Campinas",
        state="SP",
        website=f"https://{domain}" if domain else None,
        domain=domain,
        source_provider="google_places",
        source_url="https://maps.example/lead",
        lead_source_type=LeadSourceType.GOOGLE_PLACES,
    )


def test_exclusion_rule_service_blocks_domain_subdomains_and_records_reason(db_session) -> None:
    organization = _organization()
    db_session.add(organization)
    db_session.flush()
    service = ExclusionRuleService(db_session, organization_id=organization.id)
    rule = service.create_rule(
        rule_type="domain",
        pattern="https://www.grande-rede.com.br/lojas",
        reason="Known large chain",
    )
    lead = _lead("Rede Grande Campinas", organization=organization, domain="loja.grande-rede.com.br")
    db_session.add(lead)
    db_session.flush()

    match = service.apply_to_lead(lead)

    assert match is not None
    assert lead.is_blocked is True
    assert lead.blocked_rule_id == rule.id
    assert lead.blocked_reason == "Known large chain"
    assert lead.blocked_at is not None


def test_discovery_upsert_applies_active_exclusion_rules(db_session) -> None:
    organization = _organization()
    db_session.add(organization)
    db_session.flush()
    ExclusionRuleService(db_session, organization_id=organization.id).create_rule(
        rule_type="business_name_contains",
        pattern="Mega Chain",
        reason="Large chain name",
    )
    repository = LeadRepository(db_session, organization_id=organization.id)

    lead, created = repository.upsert_from_discovery(_candidate("Mega Chain Campinas"))

    assert created is True
    assert lead.is_blocked is True
    assert lead.blocked_reason == "Large chain name"


def test_repository_excludes_blocked_leads_by_default_and_supports_blocked_filter(db_session) -> None:
    organization = _organization()
    blocked = _lead(
        "Blocked Chain",
        organization=organization,
        is_blocked=True,
        blocked_reason="Known large chain",
    )
    allowed = _lead("Allowed Loja", organization=organization)
    db_session.add_all([organization, blocked, allowed])
    db_session.commit()

    repository = LeadRepository(db_session, organization_id=organization.id)

    assert [lead.business_name for lead in repository.list_all_leads()] == ["Allowed Loja"]
    assert {lead.business_name for lead in repository.list_all_leads(LeadListFilters(blocked="include"))} == {
        "Allowed Loja",
        "Blocked Chain",
    }
    assert [lead.business_name for lead in repository.list_all_leads(LeadListFilters(blocked="only"))] == [
        "Blocked Chain",
    ]
    assert [lead.id for lead in repository.list_export_leads_by_ids([blocked.id, allowed.id])] == [allowed.id]
    assert [lead.id for lead in repository.list_export_leads_by_ids([blocked.id, allowed.id], blocked="include")] == [
        blocked.id,
        allowed.id,
    ]


def test_export_excludes_blocked_leads_by_default_for_filtered_and_explicit_scopes(db_session) -> None:
    organization = _organization("default")
    blocked = _lead(
        "Blocked Export",
        organization=organization,
        is_blocked=True,
        blocked_reason="Known large chain",
    )
    allowed = _lead("Allowed Export", organization=organization)
    db_session.add_all([organization, blocked, allowed])
    db_session.commit()

    service = ExcelExportService(db_session)
    _, filtered_payload = service.build_workbook(LeadListFilters())
    _, explicit_payload = service.build_workbook(LeadListFilters(), lead_ids=[blocked.id, allowed.id])

    assert _exported_names(filtered_payload) == ["Allowed Export"]
    assert _exported_names(explicit_payload) == ["Allowed Export"]


def test_reapply_unblocks_leads_when_rule_is_deactivated(db_session) -> None:
    organization = _organization()
    db_session.add(organization)
    db_session.flush()
    service = ExclusionRuleService(db_session, organization_id=organization.id)
    rule = service.create_rule(
        rule_type="provider_category",
        pattern="hipermercado",
        reason="Provider category excluded",
    )
    lead = _lead("Mercado Grande", organization=organization, category="Hipermercado")
    db_session.add(lead)
    db_session.flush()
    service.apply_to_lead(lead)
    rule.is_active = False
    db_session.flush()

    summary = service.reapply_all()

    assert summary.evaluated == 1
    assert summary.unblocked == 1
    assert lead.is_blocked is False
    assert lead.blocked_reason is None
    assert lead.blocked_rule_id is None


def test_workshops_are_allowed_without_explicit_exclusion_rule(db_session) -> None:
    organization = _organization()
    db_session.add(organization)
    db_session.flush()

    candidate = _candidate("Oficina Motor Sul", category="oficina mecanica")
    service = ExclusionRuleService(db_session, organization_id=organization.id)
    repository = LeadRepository(db_session, organization_id=organization.id)

    assert service.evaluate_candidate(candidate) is None

    lead, created = repository.upsert_from_discovery(candidate)

    assert created is True
    assert lead.is_blocked is False
    assert lead.blocked_reason is None


def test_explicit_provider_category_rule_can_block_workshops(db_session) -> None:
    organization = _organization()
    db_session.add(organization)
    db_session.flush()
    service = ExclusionRuleService(db_session, organization_id=organization.id)
    repository = LeadRepository(db_session, organization_id=organization.id)
    service.create_rule(
        rule_type="provider_category",
        pattern="oficina mecanica",
        reason="Configured automotive exclusion",
    )

    lead, created = repository.upsert_from_discovery(
        _candidate("Oficina Motor Sul", category="oficina mecanica")
    )

    assert created is True
    assert lead.is_blocked is True
    assert lead.blocked_reason == "Configured automotive exclusion"


def test_import_exclusion_rules_script_imports_csv_and_reapplies(monkeypatch, db_session, capsys) -> None:
    organization = _organization("default")
    lead = _lead("Rede Nacional Campinas", organization=organization)
    db_session.add_all([organization, lead])
    db_session.commit()
    csv_path = Path(".pytest_cache") / "exclusion_rules_import_test.csv"
    csv_path.parent.mkdir(exist_ok=True)
    csv_path.write_text(
        "rule_type,pattern,reason,is_active\n"
        "business_name_contains,Rede Nacional,Known large chain,true\n",
        encoding="utf-8",
    )

    monkeypatch.setattr(import_exclusion_rules, "init_db", lambda: None)
    monkeypatch.setattr(import_exclusion_rules, "session_scope", lambda: _SessionScope(db_session))
    monkeypatch.setattr(sys, "argv", ["import_exclusion_rules.py", str(csv_path), "--reapply"])

    import_exclusion_rules.main()

    output = capsys.readouterr().out
    assert "Imported or updated 1 exclusion rule(s)." in output
    assert "Reapplied rules:" in output
    assert lead.is_blocked is True
    assert lead.blocked_reason == "Known large chain"


def test_create_exclusion_rule_endpoint_reapplies_existing_leads(client, db_session) -> None:
    organization = _organization("default")
    lead = _lead("Inline Block Loja", organization=organization)
    db_session.add_all([organization, lead])
    db_session.commit()

    response = client.post(
        "/exclusion-rules",
        json={
            "rule_type": "exact_name",
            "pattern": "Inline Block Loja",
            "reason": "Blocked during discovery",
        },
    )

    assert response.status_code == 200
    payload = response.json()
    assert payload["rule"]["rule_type"] == "exact_name"
    assert payload["rule"]["pattern"] == "Inline Block Loja"
    assert payload["rule"]["reason"] == "Blocked during discovery"
    assert payload["reapply_summary"]["evaluated"] == 1
    assert payload["reapply_summary"]["blocked"] == 1
    db_session.refresh(lead)
    assert lead.is_blocked is True
    assert lead.blocked_reason == "Blocked during discovery"


def test_create_exclusion_rule_endpoint_is_inline_only(client) -> None:
    response = client.post(
        "/exclusion-rules",
        json={
            "rule_type": "provider_category",
            "pattern": "hipermercado",
            "reason": "Not exposed in inline flow",
        },
    )

    assert response.status_code == 422


def _exported_names(payload: bytes) -> list[str]:
    workbook = load_workbook(BytesIO(payload))
    return [
        row[0].value
        for row in workbook["Empresas"].iter_rows(min_row=2, max_col=1)
        if row[0].value
    ]


class _SessionScope:
    def __init__(self, db_session) -> None:
        self.db_session = db_session

    def __enter__(self):
        return self.db_session

    def __exit__(self, exc_type, _exc, _traceback) -> bool:
        if exc_type is None:
            self.db_session.commit()
        else:
            self.db_session.rollback()
        return False
