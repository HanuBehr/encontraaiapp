from __future__ import annotations

from datetime import datetime, timezone
from io import BytesIO

from openpyxl import load_workbook

from app.api.routes import discovery as discovery_routes, leads as leads_routes
from app.enums import ImportBatchStatus, ImportBatchType, LeadSourceType, LeadStatus
from app.models.assignment_rule import AssignmentRule
from app.models.import_batch import ImportBatch
from app.models.lead import Lead
from app.models.market_taxonomy import MarketSegment, MarketSubsegment
from app.models.organization import Organization
from app.models.raw_discovery_record import RawDiscoveryRecord
from app.models.sales_region import SalesRegion
from app.models.sales_rep import SalesRep
from app.schemas.discovery import (
    DiscoveryPreviewEnrichmentMetadata,
    DiscoveryPreviewEnrichmentResponse,
    DiscoveryPreviewEnrichmentSummary,
    DiscoveryPreviewResponse,
    DiscoveryPreviewWebsiteRecoveryResponse,
    DiscoveryPreviewWebsiteRecoverySummary,
)
from app.schemas.lead import EnrichmentRunResult, LeadBatchEnrichmentResponse, LeadBatchEnrichmentSummary
from app.services.normalization import normalize_business_name
from app.services.providers.google_places import GooglePlacesProviderError


def _default_org() -> Organization:
    return Organization(slug="default", name="Default Organization", display_name="Default Organization")


def _lead(
    name: str,
    *,
    organization: Organization,
    city: str = "Campinas",
    state: str = "SP",
    category: str = "loja de tintas",
    instagram: str | None = None,
    status: LeadStatus = LeadStatus.NEW,
    is_blocked: bool = False,
    blocked_reason: str | None = None,
) -> Lead:
    return Lead(
        organization=organization,
        business_name=name,
        normalized_business_name=normalize_business_name(name) or name.lower(),
        category=category,
        city=city,
        state=state,
        instagram=instagram,
        lead_source_type=LeadSourceType.GOOGLE_PLACES,
        status=status,
        is_blocked=is_blocked,
        blocked_reason=blocked_reason,
    )


def _workbook_names(payload: bytes) -> list[str]:
    workbook = load_workbook(BytesIO(payload))
    return [
        row[0].value
        for row in workbook["Leads"].iter_rows(min_row=2, max_col=1)
        if row[0].value
    ]


def test_list_leads_supports_state_instagram_and_sort(client, db_session) -> None:
    organization = _default_org()
    beta = _lead(
        "Beta Tintas",
        organization=organization,
        state="SP",
        instagram="https://instagram.com/beta",
    )
    alpha = _lead(
        "Alpha Tintas",
        organization=organization,
        state="SP",
        instagram="https://instagram.com/alpha",
    )
    rio = _lead(
        "Rio Tintas",
        organization=organization,
        city="Rio de Janeiro",
        state="RJ",
        instagram="https://instagram.com/rio",
    )
    db_session.add_all([organization, beta, alpha, rio])
    db_session.commit()

    response = client.get(
        "/leads",
        params={
            "state": "SP",
            "has_instagram": True,
            "sort_by": "business_name",
            "sort_dir": "asc",
        },
    )

    assert response.status_code == 200
    payload = response.json()
    assert payload["total"] == 2
    assert [item["business_name"] for item in payload["items"]] == ["Alpha Tintas", "Beta Tintas"]
    assert [item["instagram"] for item in payload["items"]] == [
        "https://instagram.com/alpha",
        "https://instagram.com/beta",
    ]


def test_lead_options_return_v2_filter_data(client, db_session) -> None:
    organization = _default_org()
    sales_rep = SalesRep(organization=organization, name="Ana Comercial", email="ana@example.com")
    region = SalesRegion(
        organization=organization,
        name="Campinas",
        region_type="mesoregion",
        state="SP",
        code="sp-campinas",
    )
    segment = MarketSegment(organization=organization, key="varejo", name="Varejo", sort_order=1)
    subsegment = MarketSubsegment(
        organization=organization,
        segment=segment,
        key="loja_de_tintas",
        name="Loja de Tintas",
        sort_order=1,
    )
    lead = _lead("Casa Campinas", organization=organization, status=LeadStatus.REVIEWED)
    db_session.add_all([organization, sales_rep, region, segment, subsegment, lead])
    db_session.commit()

    response = client.get("/leads/options")

    assert response.status_code == 200
    payload = response.json()
    assert "Campinas" in payload["cities"]
    assert "SP" in payload["states"]
    assert "reviewed" in payload["statuses"]
    assert "ideal_sme" in payload["target_fit_values"]
    assert "varejo" in payload["trade_type_values"]
    assert payload["assigned_reps"] == [{"id": sales_rep.id, "name": "Ana Comercial"}]
    assert payload["sales_regions"][0]["code"] == "sp-campinas"
    assert payload["market_segments"] == [{"id": segment.id, "name": "Varejo", "key": "varejo"}]
    assert payload["market_subsegments"] == [
        {
            "id": subsegment.id,
            "name": "Loja de Tintas",
            "key": "loja_de_tintas",
            "market_segment_id": segment.id,
        }
    ]


def test_latest_import_batch_endpoint_and_scope_resolver(client, db_session) -> None:
    organization = _default_org()
    old_lead = _lead("Old Import Lead", organization=organization)
    new_lead = _lead("New Import Lead", organization=organization)
    old_batch = ImportBatch(
        batch_type=ImportBatchType.DISCOVERY,
        status=ImportBatchStatus.COMPLETED,
        source_provider="google_places",
        source_query="old",
        record_count=1,
        completed_at=datetime(2026, 1, 1, tzinfo=timezone.utc),
    )
    new_batch = ImportBatch(
        batch_type=ImportBatchType.DISCOVERY,
        status=ImportBatchStatus.COMPLETED,
        source_provider="google_places",
        source_query="new",
        record_count=1,
        completed_at=datetime(2026, 1, 2, tzinfo=timezone.utc),
    )
    db_session.add_all([organization, old_lead, new_lead, old_batch, new_batch])
    db_session.flush()
    db_session.add_all(
        [
            RawDiscoveryRecord(
                import_batch_id=old_batch.id,
                lead_id=old_lead.id,
                provider="google_places",
                payload_json={},
            ),
            RawDiscoveryRecord(
                import_batch_id=new_batch.id,
                lead_id=new_lead.id,
                provider="google_places",
                payload_json={},
            ),
        ]
    )
    db_session.commit()

    latest_response = client.get("/leads/import-batches/latest")
    scope_response = client.post("/leads/resolve-scope", json={"latest_import_batch": True})

    assert latest_response.status_code == 200
    assert latest_response.json()["id"] == new_batch.id
    assert latest_response.json()["lead_ids"] == [new_lead.id]
    assert scope_response.status_code == 200
    assert scope_response.json()["scope_type"] == "latest_import_batch"
    assert scope_response.json()["lead_ids"] == [new_lead.id]
    assert scope_response.json()["import_batch"]["id"] == new_batch.id


def test_import_batch_id_filters_leads_scope_and_export(client, db_session) -> None:
    organization = _default_org()
    allowed = _lead("Target Batch Lead", organization=organization)
    blocked = _lead(
        "Blocked Batch Lead",
        organization=organization,
        is_blocked=True,
        blocked_reason="Known chain",
    )
    outside = _lead("Outside Batch Lead", organization=organization)
    batch = ImportBatch(
        batch_type=ImportBatchType.DISCOVERY,
        status=ImportBatchStatus.COMPLETED,
        source_provider="google_places",
        source_query="target batch",
        record_count=2,
        completed_at=datetime(2026, 1, 4, tzinfo=timezone.utc),
    )
    outside_batch = ImportBatch(
        batch_type=ImportBatchType.DISCOVERY,
        status=ImportBatchStatus.COMPLETED,
        source_provider="google_places",
        source_query="outside batch",
        record_count=1,
        completed_at=datetime(2026, 1, 5, tzinfo=timezone.utc),
    )
    db_session.add_all([organization, allowed, blocked, outside, batch, outside_batch])
    db_session.flush()
    db_session.add_all(
        [
            RawDiscoveryRecord(
                import_batch_id=batch.id,
                lead_id=allowed.id,
                provider="google_places",
                payload_json={},
            ),
            RawDiscoveryRecord(
                import_batch_id=batch.id,
                lead_id=blocked.id,
                provider="google_places",
                payload_json={},
            ),
            RawDiscoveryRecord(
                import_batch_id=outside_batch.id,
                lead_id=outside.id,
                provider="google_places",
                payload_json={},
            ),
        ]
    )
    db_session.commit()

    default_response = client.get("/leads", params={"import_batch_id": batch.id})
    include_response = client.get(
        "/leads",
        params={
            "import_batch_id": batch.id,
            "blocked": "include",
            "sort_by": "business_name",
            "sort_dir": "asc",
        },
    )
    scope_response = client.post(
        "/leads/resolve-scope",
        json={
            "filters": {
                "import_batch_id": batch.id,
                "blocked": "include",
                "sort_by": "business_name",
                "sort_dir": "asc",
            }
        },
    )
    export_response = client.post(
        "/exports/excel",
        json={
            "filters": {
                "import_batch_id": batch.id,
                "blocked": "include",
                "sort_by": "business_name",
                "sort_dir": "asc",
            }
        },
    )

    assert default_response.status_code == 200
    assert default_response.json()["total"] == 1
    assert [item["business_name"] for item in default_response.json()["items"]] == ["Target Batch Lead"]
    assert include_response.status_code == 200
    assert [item["business_name"] for item in include_response.json()["items"]] == [
        "Blocked Batch Lead",
        "Target Batch Lead",
    ]
    assert scope_response.status_code == 200
    assert scope_response.json()["scope_label"] == f"Import batch #{batch.id} filtered set"
    assert scope_response.json()["lead_ids"] == [blocked.id, allowed.id]
    assert scope_response.json()["import_batch"]["id"] == batch.id
    assert export_response.status_code == 200
    assert _workbook_names(export_response.content) == ["Blocked Batch Lead", "Target Batch Lead"]


def test_discovery_enrich_preview_endpoint_returns_preview_summary(client, monkeypatch) -> None:
    class FakeDiscoveryService:
        def __init__(self, *args, **kwargs) -> None:
            pass

        def enrich_preview(self, *, preview, client_result_ids, skip_blocked):
            assert client_result_ids == ["preview-1"]
            assert skip_blocked is True
            return DiscoveryPreviewEnrichmentResponse(
                preview=DiscoveryPreviewResponse.model_validate(
                    {
                        "provider": "google_places",
                        "resolved_location": {
                            "label": "Campinas, SP",
                            "latitude": -22.9056,
                            "longitude": -47.0608,
                        },
                        "total_provider_results": 1,
                        "items": [
                            {
                                "client_result_id": "preview-1",
                                "search_term": "loja de tintas",
                                "provider_record_id": "place-preview-1",
                                "source_url": "https://maps.google.com/?q=Preview",
                                "raw_payload": {},
                                "candidate": {
                                    "business_name": "Preview Tintas",
                                    "normalized_business_name": "preview tintas",
                                    "category": "loja de tintas",
                                    "city": "Campinas",
                                    "state": "SP",
                                    "website": "https://preview.example.com",
                                    "domain": "preview.example.com",
                                    "email": "vendas@preview.example.com",
                                    "instagram": "https://instagram.com/preview",
                                    "phone": None,
                                    "whatsapp": None,
                                    "address": None,
                                    "neighborhood": None,
                                    "postal_code": None,
                                    "latitude": None,
                                    "longitude": None,
                                    "google_maps_url": None,
                                    "google_place_id": None,
                                    "source_provider": "google_places",
                                    "source_url": "https://maps.google.com/?q=Preview",
                                    "lead_source_type": "google_places",
                                },
                                "exclusion": {
                                    "is_blocked": False,
                                    "rule_id": None,
                                    "rule_type": None,
                                    "pattern": None,
                                    "reason": None,
                                },
                                "enrichment": DiscoveryPreviewEnrichmentMetadata(
                                    email_found=True,
                                    instagram_found=True,
                                ).model_dump(mode="json"),
                            }
                        ],
                    }
                ),
                summary=DiscoveryPreviewEnrichmentSummary(
                    requested=1,
                    processed=1,
                    success_count=1,
                    emails_found=1,
                    instagrams_found=1,
                    contact_forms_found=0,
                    no_email_found=0,
                    skipped_no_website=0,
                    blocked_after_enrichment=0,
                    errors=0,
                ),
            )

    monkeypatch.setattr(discovery_routes, "DiscoveryService", FakeDiscoveryService)

    response = client.post(
        "/discovery/enrich-preview",
        json={
            "preview": {
                "provider": "google_places",
                "resolved_location": {
                    "label": "Campinas, SP",
                    "latitude": -22.9056,
                    "longitude": -47.0608,
                },
                "total_provider_results": 1,
                "items": [
                    {
                        "client_result_id": "preview-1",
                        "search_term": "loja de tintas",
                        "provider_record_id": "place-preview-1",
                        "source_url": "https://maps.google.com/?q=Preview",
                        "raw_payload": {},
                        "candidate": {
                            "business_name": "Preview Tintas",
                            "normalized_business_name": "preview tintas",
                            "category": "loja de tintas",
                            "city": "Campinas",
                            "state": "SP",
                            "website": "https://preview.example.com",
                            "domain": None,
                            "email": None,
                            "instagram": None,
                            "phone": None,
                            "whatsapp": None,
                            "address": None,
                            "neighborhood": None,
                            "postal_code": None,
                            "latitude": None,
                            "longitude": None,
                            "google_maps_url": None,
                            "google_place_id": None,
                            "source_provider": "google_places",
                            "source_url": "https://maps.google.com/?q=Preview",
                            "lead_source_type": "google_places",
                        },
                        "exclusion": {
                            "is_blocked": False,
                            "rule_id": None,
                            "rule_type": None,
                            "pattern": None,
                            "reason": None,
                        },
                        "enrichment": None,
                    }
                ],
            },
            "client_result_ids": ["preview-1"],
            "skip_blocked": True,
        },
    )

    assert response.status_code == 200
    payload = response.json()
    assert payload["summary"]["requested"] == 1
    assert payload["summary"]["success_count"] == 1
    assert payload["summary"]["emails_found"] == 1
    assert payload["preview"]["items"][0]["candidate"]["email"] == "vendas@preview.example.com"
    assert payload["preview"]["items"][0]["enrichment"]["email_found"] is True


def test_discovery_preview_endpoint_returns_503_for_google_places_provider_error(client, monkeypatch) -> None:
    def raise_provider_error(self, payload):
        raise GooglePlacesProviderError(
            "Google Places request failed due to an upstream SSL/network error. Retry shortly.",
            status_code=503,
        )

    monkeypatch.setattr(discovery_routes.DiscoveryService, "preview", raise_provider_error)

    response = client.post(
        "/discovery/preview",
        json={
            "search_terms": ["loja de tintas"],
            "location_query": "Campinas, SP",
            "radius_m": 2500,
            "max_results_per_term": 5,
        },
    )

    assert response.status_code == 503
    assert response.json() == {
        "detail": "Google Places request failed due to an upstream SSL/network error. Retry shortly."
    }


def test_discovery_preview_endpoint_returns_503_when_google_api_key_is_missing(client) -> None:
    response = client.post(
        "/discovery/preview",
        json={
            "search_terms": ["dentistas"],
            "location_query": "São Paulo, SP",
            "radius_m": 2500,
            "max_results_per_term": 5,
        },
    )

    assert response.status_code == 503
    assert response.json() == {
        "detail": "GOOGLE_API_KEY must be configured to use location-based discovery."
    }


def test_discovery_recover_websites_endpoint_returns_preview_and_summary(client, monkeypatch) -> None:
    class FakeDiscoveryService:
        def __init__(self, *args, **kwargs) -> None:
            return None

        def recover_preview_websites(self, *, preview, client_result_ids, skip_blocked):
            return DiscoveryPreviewWebsiteRecoveryResponse(
                preview=DiscoveryPreviewResponse.model_validate(
                    {
                        "provider": "google_places",
                        "resolved_location": {
                            "label": "Campinas, SP",
                            "latitude": -22.9056,
                            "longitude": -47.0608,
                        },
                        "total_provider_results": 1,
                        "items": [
                            {
                                "client_result_id": "preview-1",
                                "search_term": "loja de tintas",
                                "provider_record_id": "place-preview-1",
                                "source_url": "https://maps.google.com/?q=Preview",
                                "raw_payload": {
                                    "id": "place-preview-1",
                                    "websiteUri": "https://preview.example.com",
                                },
                                "candidate": {
                                    "business_name": "Preview Tintas",
                                    "normalized_business_name": "preview tintas",
                                    "category": "loja de tintas",
                                    "city": "Campinas",
                                    "state": "SP",
                                    "website": "https://preview.example.com",
                                    "domain": "preview.example.com",
                                    "email": None,
                                    "instagram": None,
                                    "phone": None,
                                    "whatsapp": None,
                                    "address": None,
                                    "neighborhood": None,
                                    "postal_code": None,
                                    "latitude": None,
                                    "longitude": None,
                                    "google_maps_url": None,
                                    "google_place_id": "place-preview-1",
                                    "source_provider": "google_places",
                                    "source_url": "https://maps.google.com/?q=Preview",
                                    "lead_source_type": "google_places",
                                },
                                "exclusion": {
                                    "is_blocked": False,
                                    "rule_id": None,
                                    "rule_type": None,
                                    "pattern": None,
                                    "reason": None,
                                },
                                "enrichment": None,
                            }
                        ],
                    }
                ),
                summary=DiscoveryPreviewWebsiteRecoverySummary(
                    requested=1,
                    processed=1,
                    recovered_count=1,
                    no_website_found=0,
                    skipped_existing_website=0,
                    skipped_missing_place_id=0,
                    skipped_blocked=0,
                    blocked_after_recovery=0,
                    errors=0,
                ),
            )

    monkeypatch.setattr(discovery_routes, "DiscoveryService", FakeDiscoveryService)

    response = client.post(
        "/discovery/recover-websites",
        json={
            "preview": {
                "provider": "google_places",
                "resolved_location": {
                    "label": "Campinas, SP",
                    "latitude": -22.9056,
                    "longitude": -47.0608,
                },
                "total_provider_results": 1,
                "items": [
                    {
                        "client_result_id": "preview-1",
                        "search_term": "loja de tintas",
                        "provider_record_id": "place-preview-1",
                        "source_url": "https://maps.google.com/?q=Preview",
                        "raw_payload": {},
                        "candidate": {
                            "business_name": "Preview Tintas",
                            "normalized_business_name": "preview tintas",
                            "category": "loja de tintas",
                            "city": "Campinas",
                            "state": "SP",
                            "website": None,
                            "domain": None,
                            "email": None,
                            "instagram": None,
                            "phone": None,
                            "whatsapp": None,
                            "address": None,
                            "neighborhood": None,
                            "postal_code": None,
                            "latitude": None,
                            "longitude": None,
                            "google_maps_url": None,
                            "google_place_id": "place-preview-1",
                            "source_provider": "google_places",
                            "source_url": "https://maps.google.com/?q=Preview",
                            "lead_source_type": "google_places",
                        },
                        "exclusion": {
                            "is_blocked": False,
                            "rule_id": None,
                            "rule_type": None,
                            "pattern": None,
                            "reason": None,
                        },
                        "enrichment": None,
                    }
                ],
            },
            "client_result_ids": ["preview-1"],
            "skip_blocked": True,
        },
    )

    assert response.status_code == 200
    payload = response.json()
    assert payload["summary"]["requested"] == 1
    assert payload["summary"]["recovered_count"] == 1
    assert payload["preview"]["items"][0]["candidate"]["website"] == "https://preview.example.com"
    assert payload["preview"]["items"][0]["candidate"]["domain"] == "preview.example.com"


def test_discovery_enrich_preview_endpoint_enforces_max_25(client) -> None:
    response = client.post(
        "/discovery/enrich-preview",
        json={
            "preview": {
                "provider": "google_places",
                "resolved_location": {
                    "label": "Campinas, SP",
                    "latitude": -22.9056,
                    "longitude": -47.0608,
                },
                "total_provider_results": 1,
                "items": [],
            },
            "client_result_ids": [f"preview-{index}" for index in range(26)],
            "skip_blocked": True,
        },
    )

    assert response.status_code == 422


def test_discovery_enrich_preview_endpoint_handles_exclusion_recheck_failures(client, monkeypatch) -> None:
    def fake_enrich_preview_candidate(self, candidate):
        return (
            candidate.model_copy(
                update={
                    "email": "vendas@preview.example.com",
                    "domain": "preview.example.com",
                }
            ),
            DiscoveryPreviewEnrichmentMetadata(email_found=True),
        )

    def raise_exclusion_error(self, candidate):
        raise RuntimeError("unexpected candidate payload")

    monkeypatch.setattr(
        "app.services.discovery.EnrichmentService.enrich_preview_candidate",
        fake_enrich_preview_candidate,
    )
    monkeypatch.setattr(
        "app.services.exclusion_rules.ExclusionRuleService.evaluate_candidate",
        raise_exclusion_error,
    )

    response = client.post(
        "/discovery/enrich-preview",
        json={
            "preview": {
                "provider": "google_places",
                "resolved_location": {
                    "label": "Campinas, SP",
                    "latitude": -22.9056,
                    "longitude": -47.0608,
                },
                "total_provider_results": 1,
                "items": [
                    {
                        "client_result_id": "preview-1",
                        "search_term": "loja de tintas",
                        "provider_record_id": "place-preview-1",
                        "source_url": "https://maps.google.com/?q=Preview",
                        "raw_payload": {},
                        "candidate": {
                            "business_name": "Preview Tintas",
                            "normalized_business_name": "preview tintas",
                            "category": "loja de tintas",
                            "city": "Campinas",
                            "state": "SP",
                            "website": "https://preview.example.com",
                            "domain": None,
                            "email": None,
                            "instagram": None,
                            "phone": None,
                            "whatsapp": None,
                            "address": None,
                            "neighborhood": None,
                            "postal_code": None,
                            "latitude": None,
                            "longitude": None,
                            "google_maps_url": None,
                            "google_place_id": None,
                            "source_provider": "google_places",
                            "source_url": "https://maps.google.com/?q=Preview",
                            "lead_source_type": "google_places",
                        },
                        "exclusion": {
                            "is_blocked": False,
                            "rule_id": None,
                            "rule_type": None,
                            "pattern": None,
                            "reason": None,
                        },
                        "enrichment": None,
                    }
                ],
            },
            "client_result_ids": ["preview-1"],
            "skip_blocked": True,
        },
    )

    assert response.status_code == 200
    payload = response.json()
    assert payload["summary"]["requested"] == 1
    assert payload["summary"]["processed"] == 1
    assert payload["summary"]["success_count"] == 1
    assert payload["summary"]["errors"] == 0
    assert payload["preview"]["items"][0]["candidate"]["email"] == "vendas@preview.example.com"
    assert payload["preview"]["items"][0]["exclusion"]["is_blocked"] is False


def test_discovery_enrich_preview_endpoint_returns_200_when_one_selected_row_fails(client, monkeypatch) -> None:
    def fake_enrich_preview_candidate(self, candidate):
        if candidate.business_name == "Preview Ruim":
            return candidate, object()
        return (
            candidate.model_copy(
                update={"email": f"{candidate.normalized_business_name.replace(' ', '')}@example.com"}
            ),
            DiscoveryPreviewEnrichmentMetadata(email_found=True),
        )

    monkeypatch.setattr(
        "app.services.discovery.EnrichmentService.enrich_preview_candidate",
        fake_enrich_preview_candidate,
    )

    response = client.post(
        "/discovery/enrich-preview",
        json={
            "preview": {
                "provider": "google_places",
                "resolved_location": {
                    "label": "Campinas, SP",
                    "latitude": -22.9056,
                    "longitude": -47.0608,
                },
                "total_provider_results": 3,
                "items": [
                    {
                        "client_result_id": "preview-good-1",
                        "search_term": "loja de tintas",
                        "provider_record_id": "place-preview-good-1",
                        "source_url": "https://maps.google.com/?q=Preview+Boa+1",
                        "raw_payload": {},
                        "candidate": {
                            "business_name": "Preview Boa 1",
                            "normalized_business_name": "preview boa 1",
                            "category": "loja de tintas",
                            "city": "Campinas",
                            "state": "SP",
                            "website": "https://preview-boa-1.example.com",
                            "domain": None,
                            "email": None,
                            "instagram": None,
                            "phone": None,
                            "whatsapp": None,
                            "address": None,
                            "neighborhood": None,
                            "postal_code": None,
                            "latitude": None,
                            "longitude": None,
                            "google_maps_url": None,
                            "google_place_id": None,
                            "source_provider": "google_places",
                            "source_url": "https://maps.google.com/?q=Preview+Boa+1",
                            "lead_source_type": "google_places",
                        },
                        "exclusion": {
                            "is_blocked": False,
                            "rule_id": None,
                            "rule_type": None,
                            "pattern": None,
                            "reason": None,
                        },
                        "enrichment": None,
                    },
                    {
                        "client_result_id": "preview-bad-1",
                        "search_term": "loja de tintas",
                        "provider_record_id": "place-preview-bad-1",
                        "source_url": "https://maps.google.com/?q=Preview+Ruim",
                        "raw_payload": {},
                        "candidate": {
                            "business_name": "Preview Ruim",
                            "normalized_business_name": "preview ruim",
                            "category": "loja de tintas",
                            "city": "Campinas",
                            "state": "SP",
                            "website": "https://preview-ruim.example.com",
                            "domain": None,
                            "email": None,
                            "instagram": None,
                            "phone": None,
                            "whatsapp": None,
                            "address": None,
                            "neighborhood": None,
                            "postal_code": None,
                            "latitude": None,
                            "longitude": None,
                            "google_maps_url": None,
                            "google_place_id": None,
                            "source_provider": "google_places",
                            "source_url": "https://maps.google.com/?q=Preview+Ruim",
                            "lead_source_type": "google_places",
                        },
                        "exclusion": {
                            "is_blocked": False,
                            "rule_id": None,
                            "rule_type": None,
                            "pattern": None,
                            "reason": None,
                        },
                        "enrichment": None,
                    },
                    {
                        "client_result_id": "preview-good-2",
                        "search_term": "loja de tintas",
                        "provider_record_id": "place-preview-good-2",
                        "source_url": "https://maps.google.com/?q=Preview+Boa+2",
                        "raw_payload": {},
                        "candidate": {
                            "business_name": "Preview Boa 2",
                            "normalized_business_name": "preview boa 2",
                            "category": "loja de tintas",
                            "city": "Campinas",
                            "state": "SP",
                            "website": "https://preview-boa-2.example.com",
                            "domain": None,
                            "email": None,
                            "instagram": None,
                            "phone": None,
                            "whatsapp": None,
                            "address": None,
                            "neighborhood": None,
                            "postal_code": None,
                            "latitude": None,
                            "longitude": None,
                            "google_maps_url": None,
                            "google_place_id": None,
                            "source_provider": "google_places",
                            "source_url": "https://maps.google.com/?q=Preview+Boa+2",
                            "lead_source_type": "google_places",
                        },
                        "exclusion": {
                            "is_blocked": False,
                            "rule_id": None,
                            "rule_type": None,
                            "pattern": None,
                            "reason": None,
                        },
                        "enrichment": None,
                    },
                ],
            },
            "client_result_ids": ["preview-good-1", "preview-bad-1", "preview-good-2"],
            "skip_blocked": True,
        },
    )

    assert response.status_code == 200
    payload = response.json()
    assert payload["summary"]["requested"] == 3
    assert payload["summary"]["processed"] == 3
    assert payload["summary"]["success_count"] == 2
    assert payload["summary"]["errors"] == 1
    assert payload["summary"]["emails_found"] == 2
    by_id = {item["client_result_id"]: item for item in payload["preview"]["items"]}
    assert by_id["preview-good-1"]["candidate"]["email"] == "previewboa1@example.com"
    assert by_id["preview-bad-1"]["enrichment"]["success"] is False
    assert by_id["preview-bad-1"]["enrichment"]["error_message"] == "'object' object has no attribute 'model_dump'"
    assert by_id["preview-good-2"]["candidate"]["email"] == "previewboa2@example.com"


def test_discovery_enrich_preview_endpoint_retries_rows_when_multi_row_request_fails(client, monkeypatch) -> None:
    original_once = discovery_routes.DiscoveryService._enrich_preview_once

    def flaky_enrich_preview_once(self, *, preview, client_result_ids, skip_blocked):
        if len(client_result_ids) > 1:
            raise RuntimeError("combined preview enrichment crash")
        client_result_id = client_result_ids[0]
        if client_result_id == "preview-bad-1":
            raise RuntimeError("single row parser crash")
        return original_once(
            self,
            preview=preview,
            client_result_ids=client_result_ids,
            skip_blocked=skip_blocked,
        )

    def fake_enrich_preview_candidate(self, candidate):
        return (
            candidate.model_copy(
                update={"email": f"{candidate.normalized_business_name.replace(' ', '')}@example.com"}
            ),
            DiscoveryPreviewEnrichmentMetadata(email_found=True),
        )

    monkeypatch.setattr(discovery_routes.DiscoveryService, "_enrich_preview_once", flaky_enrich_preview_once)
    monkeypatch.setattr(
        "app.services.discovery.EnrichmentService.enrich_preview_candidate",
        fake_enrich_preview_candidate,
    )

    response = client.post(
        "/discovery/enrich-preview",
        json={
            "preview": {
                "provider": "google_places",
                "resolved_location": {
                    "label": "Campinas, SP",
                    "latitude": -22.9056,
                    "longitude": -47.0608,
                },
                "total_provider_results": 3,
                "items": [
                    {
                        "client_result_id": "preview-good-1",
                        "search_term": "loja de tintas",
                        "provider_record_id": "place-preview-good-1",
                        "source_url": "https://maps.google.com/?q=Preview+Boa+1",
                        "raw_payload": {},
                        "candidate": {
                            "business_name": "Preview Boa 1",
                            "normalized_business_name": "preview boa 1",
                            "category": "loja de tintas",
                            "city": "Campinas",
                            "state": "SP",
                            "website": "https://preview-boa-1.example.com",
                            "domain": None,
                            "email": None,
                            "instagram": None,
                            "phone": None,
                            "whatsapp": None,
                            "address": None,
                            "neighborhood": None,
                            "postal_code": None,
                            "latitude": None,
                            "longitude": None,
                            "google_maps_url": None,
                            "google_place_id": None,
                            "source_provider": "google_places",
                            "source_url": "https://maps.google.com/?q=Preview+Boa+1",
                            "lead_source_type": "google_places",
                        },
                        "exclusion": {
                            "is_blocked": False,
                            "rule_id": None,
                            "rule_type": None,
                            "pattern": None,
                            "reason": None,
                        },
                        "enrichment": None,
                    },
                    {
                        "client_result_id": "preview-bad-1",
                        "search_term": "loja de tintas",
                        "provider_record_id": "place-preview-bad-1",
                        "source_url": "https://maps.google.com/?q=Preview+Ruim",
                        "raw_payload": {},
                        "candidate": {
                            "business_name": "Preview Ruim",
                            "normalized_business_name": "preview ruim",
                            "category": "loja de tintas",
                            "city": "Campinas",
                            "state": "SP",
                            "website": "https://preview-ruim.example.com",
                            "domain": None,
                            "email": None,
                            "instagram": None,
                            "phone": None,
                            "whatsapp": None,
                            "address": None,
                            "neighborhood": None,
                            "postal_code": None,
                            "latitude": None,
                            "longitude": None,
                            "google_maps_url": None,
                            "google_place_id": None,
                            "source_provider": "google_places",
                            "source_url": "https://maps.google.com/?q=Preview+Ruim",
                            "lead_source_type": "google_places",
                        },
                        "exclusion": {
                            "is_blocked": False,
                            "rule_id": None,
                            "rule_type": None,
                            "pattern": None,
                            "reason": None,
                        },
                        "enrichment": None,
                    },
                    {
                        "client_result_id": "preview-good-2",
                        "search_term": "loja de tintas",
                        "provider_record_id": "place-preview-good-2",
                        "source_url": "https://maps.google.com/?q=Preview+Boa+2",
                        "raw_payload": {},
                        "candidate": {
                            "business_name": "Preview Boa 2",
                            "normalized_business_name": "preview boa 2",
                            "category": "loja de tintas",
                            "city": "Campinas",
                            "state": "SP",
                            "website": "https://preview-boa-2.example.com",
                            "domain": None,
                            "email": None,
                            "instagram": None,
                            "phone": None,
                            "whatsapp": None,
                            "address": None,
                            "neighborhood": None,
                            "postal_code": None,
                            "latitude": None,
                            "longitude": None,
                            "google_maps_url": None,
                            "google_place_id": None,
                            "source_provider": "google_places",
                            "source_url": "https://maps.google.com/?q=Preview+Boa+2",
                            "lead_source_type": "google_places",
                        },
                        "exclusion": {
                            "is_blocked": False,
                            "rule_id": None,
                            "rule_type": None,
                            "pattern": None,
                            "reason": None,
                        },
                        "enrichment": None,
                    },
                ],
            },
            "client_result_ids": ["preview-good-1", "preview-bad-1", "preview-good-2"],
            "skip_blocked": True,
        },
    )

    assert response.status_code == 200
    payload = response.json()
    assert payload["summary"]["requested"] == 3
    assert payload["summary"]["processed"] == 3
    assert payload["summary"]["success_count"] == 2
    assert payload["summary"]["errors"] == 1
    assert payload["summary"]["emails_found"] == 2
    assert payload["summary"]["error_messages"] == ["Preview Ruim: single row parser crash"]
    by_id = {item["client_result_id"]: item for item in payload["preview"]["items"]}
    assert by_id["preview-good-1"]["candidate"]["email"] == "previewboa1@example.com"
    assert by_id["preview-bad-1"]["enrichment"]["success"] is False
    assert by_id["preview-bad-1"]["enrichment"]["error_message"] == "single row parser crash"
    assert by_id["preview-good-2"]["candidate"]["email"] == "previewboa2@example.com"


def test_batch_assign_wraps_existing_assignment_service(client, db_session) -> None:
    organization = _default_org()
    sales_rep = SalesRep(organization=organization, name="Ana Comercial")
    region = SalesRegion(
        organization=organization,
        name="Campinas",
        region_type="mesoregion",
        state="SP",
        cities_json=["Campinas"],
    )
    segment = MarketSegment(organization=organization, key="varejo", name="Varejo")
    subsegment = MarketSubsegment(
        organization=organization,
        segment=segment,
        key="loja_de_tintas",
        name="Loja de Tintas",
        keywords_json=["tintas"],
    )
    rule = AssignmentRule(
        organization=organization,
        name="Campinas Tintas",
        sales_region=region,
        market_segment=segment,
        market_subsegment=subsegment,
        sales_rep=sales_rep,
    )
    lead = _lead("Casa das Tintas", organization=organization, city="Campinas", state="SP")
    db_session.add_all([organization, sales_rep, region, segment, subsegment, rule, lead])
    db_session.commit()

    response = client.post("/leads/batch/assign", json={"lead_ids": [lead.id]})

    assert response.status_code == 200
    payload = response.json()
    assert payload["processed"] == 1
    assert payload["changed"] == 1
    assert payload["summary"]["scope_type"] == "lead_ids"
    assert payload["results"][0]["suggestion"]["assigned_sales_rep_id"] == sales_rep.id
    db_session.refresh(lead)
    assert lead.assigned_sales_rep_id == sales_rep.id
    assert lead.sales_region_id == region.id
    assert lead.market_subsegment_id == subsegment.id


def test_batch_enrich_returns_partial_success_response(client, monkeypatch) -> None:
    class FakeEnrichmentService:
        def __init__(self, *args, **kwargs) -> None:
            pass

        def enrich_lead_ids(self, lead_ids, *, actor, scope_label, continue_on_error):
            assert lead_ids == [10, 20]
            assert actor == "api"
            assert scope_label == "api lead ids"
            assert continue_on_error is True
            return LeadBatchEnrichmentResponse(
                processed=2,
                results=[
                    EnrichmentRunResult(
                        lead_id=10,
                        business_name="Empresa Boa",
                        pages_attempted=1,
                        pages_fetched=1,
                        contacts_added=1,
                        contacts_added_by_type={"EMAIL": 1},
                        fields_updated=["email"],
                        last_enriched_at=datetime(2026, 1, 3, tzinfo=timezone.utc),
                    ),
                    EnrichmentRunResult(
                        lead_id=20,
                        business_name="Empresa Ruim",
                        success=False,
                        error_message="parser failed on public page",
                    ),
                ],
                summary=LeadBatchEnrichmentSummary(
                    scope_label="api lead ids",
                    requested=2,
                    processed=2,
                    success_count=1,
                    contacts_added=1,
                    emails_found=1,
                    errors=1,
                    error_messages=["Lead 20: parser failed on public page"],
                    failed_lead_ids=[20],
                    pages_attempted=1,
                    pages_fetched=1,
                ),
            )

    monkeypatch.setattr(leads_routes, "EnrichmentService", FakeEnrichmentService)

    response = client.post("/leads/batch/enrich", json={"lead_ids": [10, 20]})

    assert response.status_code == 200
    payload = response.json()
    assert payload["processed"] == 2
    assert payload["summary"]["requested"] == 2
    assert payload["summary"]["processed"] == 2
    assert payload["summary"]["success_count"] == 1
    assert payload["summary"]["errors"] == 1
    assert payload["summary"]["failed_lead_ids"] == [20]
    assert payload["results"][1]["success"] is False
    assert payload["results"][1]["business_name"] == "Empresa Ruim"
    assert payload["results"][1]["error_message"] == "parser failed on public page"


def test_scoped_export_accepts_lead_ids_filters_and_latest_batch(client, db_session) -> None:
    organization = _default_org()
    alpha = _lead(
        "Alpha Export",
        organization=organization,
        state="SP",
        instagram="https://instagram.com/alpha",
    )
    beta = _lead("Beta Export", organization=organization, state="RJ")
    batch = ImportBatch(
        batch_type=ImportBatchType.DISCOVERY,
        status=ImportBatchStatus.COMPLETED,
        source_provider="google_places",
        source_query="export",
        record_count=1,
        completed_at=datetime(2026, 1, 3, tzinfo=timezone.utc),
    )
    db_session.add_all([organization, alpha, beta, batch])
    db_session.flush()
    db_session.add(
        RawDiscoveryRecord(
            import_batch_id=batch.id,
            lead_id=beta.id,
            provider="google_places",
            payload_json={},
        )
    )
    db_session.commit()

    selected_response = client.post("/exports/excel", json={"lead_ids": [alpha.id]})
    filtered_response = client.post(
        "/exports/excel",
        json={
            "filters": {
                "state": "SP",
                "has_instagram": True,
                "sort_by": "business_name",
                "sort_dir": "asc",
            }
        },
    )
    latest_response = client.post("/exports/excel", json={"latest_import_batch": True})

    assert selected_response.status_code == 200
    assert filtered_response.status_code == 200
    assert latest_response.status_code == 200
    assert _workbook_names(selected_response.content) == ["Alpha Export"]
    assert _workbook_names(filtered_response.content) == ["Alpha Export"]
    assert _workbook_names(latest_response.content) == ["Beta Export"]
