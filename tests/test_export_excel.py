from __future__ import annotations

from datetime import datetime, timezone
from io import BytesIO
from types import SimpleNamespace

from openpyxl import load_workbook

from app.enums import (
    CompanySizeFit,
    ContactType,
    ImportBatchStatus,
    ImportBatchType,
    LeadSourceType,
    LeadStatus,
    TradeType,
)
from app.models.assignment_rule import AssignmentRule
from app.models.import_batch import ImportBatch
from app.models.lead import Lead
from app.models.lead_contact import LeadContact
from app.models.market_taxonomy import MarketSegment, MarketSubsegment
from app.models.organization import Organization
from app.models.raw_discovery_record import RawDiscoveryRecord
from app.models.sales_region import SalesRegion
from app.models.sales_rep import SalesRep
from app.repositories.lead_repository import LeadRepository
from app.schemas.lead import LeadListFilters
from app.services.export_excel import ExcelExportService
from app.services.normalization import normalize_business_name


EXPECTED_LEADS_HEADERS = [
    "Nome",
    "CNPJ",
    "Razão Social",
    "Categoria",
    "Origem",
    "Usuário responsável",
    "Setor",
    "Descrição",
    "E-mail",
    "WhatsApp",
    "Telefone",
    "Celular",
    "Fax",
    "Ramal",
    "Website",
    "CEP",
    "País",
    "Estado",
    "Cidade",
    "Bairro",
    "Rua",
    "Número",
    "Complemento",
    "Produto",
    "Facebook",
    "Twitter",
    "LinkedIn",
    "Skype",
    "Instagram",
    "Ranking",
]


def _lead(name: str, *, organization: Organization, city: str = "Campinas") -> Lead:
    return Lead(
        organization=organization,
        business_name=name,
        normalized_business_name=normalize_business_name(name) or name.lower(),
        category="loja de tintas",
        city=city,
        state="SP",
        lead_source_type=LeadSourceType.GOOGLE_PLACES,
        status=LeadStatus.NEW,
    )


def _empresa_row(workbook, row_number: int = 2) -> dict[str, object]:
    headers = [cell.value for cell in workbook["Empresas"][1]]
    values = [cell.value for cell in workbook["Empresas"][row_number]]
    return dict(zip(headers, values, strict=False))


def test_excel_export_matches_empresas_template(db_session) -> None:
    organization = Organization(slug="default", name="Default Organization", display_name="Default Organization")
    region = SalesRegion(
        organization=organization,
        name="Campinas",
        region_type="mesoregion",
        state="SP",
        code="sp-campinas",
    )
    segment = MarketSegment(organization=organization, key="varejo", name="Varejo")
    subsegment = MarketSubsegment(
        organization=organization,
        segment=segment,
        key="loja_de_tintas",
        name="loja de tintas",
    )
    sales_rep = SalesRep(organization=organization, name="Vendas2")
    rule = AssignmentRule(
        organization=organization,
        name="Varejo - Vendas2 - Campinas",
        sales_region=region,
        market_segment=segment,
        sales_rep=sales_rep,
    )
    lead = Lead(
        organization=organization,
        business_name="Oficina Export",
        normalized_business_name=normalize_business_name("Oficina Export") or "oficina export",
        cnpj="37335118000180",
        legal_name="Oficina Export LTDA",
        cnpj_match_status="matched",
        category="oficina mecanica",
        address="Rua das Tintas, 100",
        city="Sao Paulo",
        state="SP",
        postal_code="01000-000",
        website="https://export.com.br",
        notes="Observação comercial",
        lead_source_type=LeadSourceType.GOOGLE_PLACES,
        status=LeadStatus.NEW,
        sales_region=region,
        market_segment=segment,
        market_subsegment=subsegment,
        assigned_sales_rep=sales_rep,
        assignment_rule=rule,
        assignment_explanation="Matched test assignment.",
        company_size_fit=CompanySizeFit.IDEAL_SME.value,
        company_size_fit_explanation="Matched SME target.",
        trade_type=TradeType.VAREJO.value,
        trade_type_explanation="Matched retail segment.",
    )
    db_session.add(lead)
    db_session.flush()
    db_session.add_all(
        [
            LeadContact(
                organization=organization,
                lead=lead,
                contact_type=ContactType.EMAIL,
                raw_value="vendas@export.com.br",
                normalized_value="vendas@export.com.br",
                confidence=0.95,
                is_primary=True,
            ),
            LeadContact(
                organization=organization,
                lead=lead,
                contact_type=ContactType.INSTAGRAM,
                raw_value="https://instagram.com/export",
                normalized_value="https://instagram.com/export",
                confidence=0.93,
                is_primary=True,
            ),
            LeadContact(
                organization=organization,
                lead=lead,
                contact_type=ContactType.WHATSAPP,
                raw_value="+55 11 99999-9999",
                normalized_value="+5511999999999",
                confidence=0.88,
                is_primary=True,
            ),
        ]
    )
    db_session.commit()
    db_session.refresh(lead)

    service = ExcelExportService(db_session)
    filename, payload = service.build_workbook(LeadListFilters())

    workbook = load_workbook(BytesIO(payload))
    lead_headers = [cell.value for cell in workbook["Empresas"][1]]
    lead_row = _empresa_row(workbook)
    metadata_rows = [row[0].value for row in workbook["Metadata"].iter_rows(min_row=2, max_col=1)]

    assert filename.endswith(".xlsx")
    assert payload[:2] == b"PK"
    assert workbook.sheetnames[0] == "Empresas"
    assert "Leads" not in workbook.sheetnames
    assert "Metadata" in workbook.sheetnames
    assert "export_timestamp_utc" in metadata_rows
    assert lead_headers == EXPECTED_LEADS_HEADERS
    assert "business_name" not in lead_headers
    assert "google_maps_url" not in lead_headers
    assert "lead_score" not in lead_headers
    assert "review_count" not in lead_headers
    assert lead_row["Nome"] == "Oficina Export"
    assert lead_row["CNPJ"] == "37335118000180"
    assert lead_row["Razão Social"] == "Oficina Export LTDA"
    assert lead_row["Categoria"] == "oficina mecanica"
    assert lead_row["Origem"] == "Google Places"
    assert lead_row["Usuário responsável"] == "Vendas2"
    assert lead_row["Setor"] == "oficina mecanica"
    assert lead_row["Descrição"] == "Observação comercial"
    assert lead_row["E-mail"] == "vendas@export.com.br"
    assert lead_row["WhatsApp"] == "+5511999999999"
    assert lead_row["Telefone"] is None
    assert lead_row["Website"] == "https://export.com.br"
    assert lead_row["CEP"] == "01000-000"
    assert lead_row["País"] == "Brasil"
    assert lead_row["Estado"] == "SP"
    assert lead_row["Cidade"] == "Sao Paulo"
    assert lead_row["Rua"] == "Rua das Tintas"
    assert lead_row["Número"] == "100"
    assert lead_row["Produto"] == "Varejo / loja de tintas"
    assert lead_row["Instagram"] == "https://instagram.com/export"
    assert lead_row["Ranking"] == 0


def test_excel_export_normalizes_client_facing_address_fields_and_filters_social_website(db_session) -> None:
    organization = Organization(slug="default", name="Default Organization", display_name="Default Organization")
    lead = Lead(
        organization=organization,
        business_name="Casa Paulista",
        normalized_business_name=normalize_business_name("Casa Paulista") or "casa paulista",
        category="materiais de construcao",
        address="Avenida Paulista, 1578 - Bela Vista, Sao Paulo - Sao Paulo, 01310-200, Brasil",
        neighborhood="Bela Vista",
        city="Sao Paulo",
        state="Sao Paulo",
        postal_code="01310-200",
        website="https://instagram.com/casapaulista",
        instagram="https://instagram.com/casapaulista",
        lead_source_type=LeadSourceType.GOOGLE_PLACES,
        status=LeadStatus.NEW,
    )
    db_session.add_all([organization, lead])
    db_session.commit()

    service = ExcelExportService(db_session)
    _, payload = service.build_workbook(LeadListFilters())

    workbook = load_workbook(BytesIO(payload))
    lead_row = _empresa_row(workbook)

    assert lead_row["Estado"] == "SP"
    assert lead_row["Cidade"] == "Sao Paulo"
    assert lead_row["Bairro"] == "Bela Vista"
    assert lead_row["Rua"] == "Avenida Paulista"
    assert lead_row["Número"] == "1578"
    assert lead_row["CEP"] == "01310-200"
    assert lead_row["Website"] is None
    assert lead_row["Instagram"] == "https://instagram.com/casapaulista"


def test_excel_export_prefers_plausible_email_and_blanks_ambiguous_number(db_session) -> None:
    organization = Organization(slug="default", name="Default Organization", display_name="Default Organization")
    lead = Lead(
        organization=organization,
        business_name="Rua Limpa",
        normalized_business_name=normalize_business_name("Rua Limpa") or "rua limpa",
        category="materiais de construcao",
        address="Endereco: Rua Alfa, 11334455 - Centro, Campinas - SP, 13010-123, Brasil",
        neighborhood="Centro",
        city="Campinas",
        state="Sao Paulo",
        postal_code="13010-123",
        lead_source_type=LeadSourceType.GOOGLE_PLACES,
        status=LeadStatus.NEW,
    )
    db_session.add_all([organization, lead])
    db_session.flush()
    db_session.add_all(
        [
            LeadContact(
                organization=organization,
                lead=lead,
                contact_type=ContactType.EMAIL,
                raw_value="alerts@sentry-next.wixpress.com",
                normalized_value="alerts@sentry-next.wixpress.com",
                confidence=0.99,
                is_primary=True,
            ),
            LeadContact(
                organization=organization,
                lead=lead,
                contact_type=ContactType.EMAIL,
                raw_value="comercial@rualimpa.com.br",
                normalized_value="comercial@rualimpa.com.br",
                confidence=0.8,
                is_primary=False,
            ),
        ]
    )
    db_session.commit()

    service = ExcelExportService(db_session)
    _, payload = service.build_workbook(LeadListFilters())

    workbook = load_workbook(BytesIO(payload))
    lead_row = _empresa_row(workbook)

    assert lead_row["Rua"] == "Rua Alfa"
    assert lead_row["Número"] is None
    assert lead_row["E-mail"] == "comercial@rualimpa.com.br"


def test_excel_export_blanks_placeholder_email_when_no_plausible_option_exists(db_session) -> None:
    organization = Organization(slug="default", name="Default Organization", display_name="Default Organization")
    lead = Lead(
        organization=organization,
        business_name="Email Placeholder",
        normalized_business_name=normalize_business_name("Email Placeholder") or "email placeholder",
        category="materiais de construcao",
        email="meu@email.com.br",
        lead_source_type=LeadSourceType.GOOGLE_PLACES,
        status=LeadStatus.NEW,
    )
    db_session.add_all([organization, lead])
    db_session.commit()

    service = ExcelExportService(db_session)
    _, payload = service.build_workbook(LeadListFilters())

    workbook = load_workbook(BytesIO(payload))
    lead_row = _empresa_row(workbook)

    assert lead_row["E-mail"] is None


def test_excel_export_strips_prefixed_numeric_garbage_from_street(db_session) -> None:
    organization = Organization(slug="default", name="Default Organization", display_name="Default Organization")
    lead = Lead(
        organization=organization,
        business_name="Casa da Pintura",
        normalized_business_name=normalize_business_name("Casa da Pintura") or "casa da pintura",
        category="materiais de construcao",
        address="8665R. Bento Goncalves, 112 - Centro, Porto Alegre - RS, 90650-001, Brasil",
        neighborhood="Centro",
        city="Porto Alegre",
        state="Rio Grande do Sul",
        postal_code="90650-001",
        lead_source_type=LeadSourceType.GOOGLE_PLACES,
        status=LeadStatus.NEW,
    )
    db_session.add_all([organization, lead])
    db_session.commit()

    service = ExcelExportService(db_session)
    _, payload = service.build_workbook(LeadListFilters())

    workbook = load_workbook(BytesIO(payload))
    lead_row = _empresa_row(workbook)

    assert lead_row["Rua"] == "R. Bento Goncalves"
    assert lead_row["Número"] == "112"
    assert lead_row["Estado"] == "RS"


def test_excel_export_with_explicit_lead_ids_preserves_scope_and_order(db_session) -> None:
    organization = Organization(slug="default", name="Default Organization", display_name="Default Organization")
    alpha = _lead("Alpha Tintas", organization=organization)
    beta = _lead("Beta Ferragens", organization=organization)
    gamma = _lead("Gamma Materiais", organization=organization)
    db_session.add_all([organization, alpha, beta, gamma])
    db_session.flush()
    db_session.add_all(
        [
            LeadContact(
                organization=organization,
                lead=beta,
                contact_type="EMAIL",
                raw_value="vendas@beta.com.br",
                normalized_value="vendas@beta.com.br",
                confidence=0.9,
                is_primary=True,
            ),
            LeadContact(
                organization=organization,
                lead=beta,
                contact_type="INSTAGRAM",
                raw_value="https://instagram.com/beta",
                normalized_value="https://instagram.com/beta",
                confidence=0.9,
                is_primary=True,
            ),
        ]
    )
    db_session.commit()

    service = ExcelExportService(db_session)
    _, payload = service.build_workbook(
        LeadListFilters(),
        lead_ids=[beta.id, alpha.id],
        scope_label="Current filtered queue",
        scope_metadata={"scope_type": "current_filtered_queue"},
    )

    workbook = load_workbook(BytesIO(payload))
    lead_rows = [
        row[0].value
        for row in workbook["Empresas"].iter_rows(min_row=2, max_col=1)
        if row[0].value
    ]
    lead_headers = [cell.value for cell in workbook["Empresas"][1]]
    first_lead_row = _empresa_row(workbook)
    metadata = {
        row[0].value: row[1].value
        for row in workbook["Metadata"].iter_rows(min_row=2, max_col=2)
        if row[0].value
    }

    assert lead_rows == ["Beta Ferragens", "Alpha Tintas"]
    assert "Gamma Materiais" not in lead_rows
    assert beta.email is None
    assert beta.instagram is None
    assert lead_headers == EXPECTED_LEADS_HEADERS
    assert first_lead_row["E-mail"] == "vendas@beta.com.br"
    assert first_lead_row["Instagram"] == "https://instagram.com/beta"
    assert metadata["scope_type"] == "current_filtered_queue"
    assert metadata["scope_label"] == "Current filtered queue"
    assert metadata["lead_count"] == 2
    assert metadata["total_lead_count"] == 2


def test_best_contact_value_matches_uppercase_raw_contact_type_strings() -> None:
    lead = SimpleNamespace(
        email=None,
        instagram=None,
        contacts=[
            SimpleNamespace(
                contact_type="EMAIL",
                raw_value="raw@example.com",
                normalized_value="normalized@example.com",
                confidence=0.5,
                is_primary=True,
                updated_at=None,
                id=1,
            ),
            SimpleNamespace(
                contact_type="INSTAGRAM",
                raw_value="https://instagram.com/raw",
                normalized_value=None,
                confidence=0.5,
                is_primary=True,
                updated_at=None,
                id=2,
            ),
        ],
    )

    assert ExcelExportService._best_contact_value(lead, ContactType.EMAIL, None) == "normalized@example.com"
    assert ExcelExportService._best_contact_value(lead, ContactType.INSTAGRAM, None) == "https://instagram.com/raw"


def test_import_batch_lead_id_resolution_is_distinct_latest_and_org_scoped(db_session) -> None:
    default_org = Organization(slug="default", name="Default Organization", display_name="Default Organization")
    other_org = Organization(slug="other", name="Other Organization", display_name="Other Organization")
    old_lead = _lead("Old Batch Lead", organization=default_org)
    new_lead = _lead("New Batch Lead", organization=default_org)
    other_lead = _lead("Other Org Lead", organization=other_org)
    old_batch = ImportBatch(
        batch_type=ImportBatchType.DISCOVERY,
        status=ImportBatchStatus.COMPLETED,
        source_provider="google_places",
        source_query="loja de tintas",
        completed_at=datetime(2026, 1, 1, tzinfo=timezone.utc),
    )
    new_batch = ImportBatch(
        batch_type=ImportBatchType.DISCOVERY,
        status=ImportBatchStatus.COMPLETED,
        source_provider="google_places",
        source_query="ferragistas",
        completed_at=datetime(2026, 1, 2, tzinfo=timezone.utc),
    )
    db_session.add_all([default_org, other_org, old_lead, new_lead, other_lead, old_batch, new_batch])
    db_session.flush()
    db_session.add_all(
        [
            RawDiscoveryRecord(
                import_batch_id=old_batch.id,
                lead_id=old_lead.id,
                provider="google_places",
                payload_json={},
            ),
            RawDiscoveryRecord(
                import_batch_id=new_batch.id,
                lead_id=new_lead.id,
                provider="google_places",
                payload_json={},
            ),
            RawDiscoveryRecord(
                import_batch_id=new_batch.id,
                lead_id=new_lead.id,
                provider="google_places",
                payload_json={},
            ),
            RawDiscoveryRecord(
                import_batch_id=new_batch.id,
                lead_id=other_lead.id,
                provider="google_places",
                payload_json={},
            ),
            RawDiscoveryRecord(
                import_batch_id=new_batch.id,
                lead_id=None,
                provider="google_places",
                payload_json={},
            ),
        ]
    )
    db_session.commit()

    repository = LeadRepository(db_session)

    assert repository.get_latest_completed_import_batch().id == new_batch.id
    assert repository.list_lead_ids_for_import_batch(new_batch.id) == [new_lead.id]
    assert repository.list_lead_ids_for_import_batch(old_batch.id) == [old_lead.id]
