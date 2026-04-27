from __future__ import annotations

from collections.abc import Iterable
from datetime import datetime, timezone
from io import BytesIO
import json

import pandas as pd
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from app.enums import ContactType
from app.repositories.lead_repository import LeadRepository
from app.schemas.lead import LeadListFilters
from app.services.normalization import (
    canonicalize_url,
    is_probable_business_website,
    is_probable_client_facing_email,
    normalize_brazilian_state,
    split_street_and_number,
)


HEADER_FILL = PatternFill(fill_type="solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
MISSING_FILL = PatternFill(fill_type="solid", fgColor="FCE4D6")
LEADS_COLUMNS = [
    "business_name",
    "category",
    "city",
    "state",
    "address",
    "street",
    "number",
    "postal_code",
    "phone",
    "whatsapp",
    "email",
    "instagram",
    "website",
    "google_maps_url",
    "source",
    "status",
    "notes",
    "neighborhood",
    "assigned_owner",
    "market_segment",
    "market_subsegment",
    "lead_score",
    "rating",
    "review_count",
]
METADATA_COLUMNS = ["metric", "value"]
SHEET_COLUMNS = {
    "Leads": LEADS_COLUMNS,
    "Metadata": METADATA_COLUMNS,
}


class ExcelExportService:
    def __init__(self, db: Session) -> None:
        self.db = db
        self.lead_repository = LeadRepository(db)

    def build_workbook(
        self,
        filters: LeadListFilters,
        *,
        lead_ids: Iterable[int] | None = None,
        scope_label: str | None = None,
        scope_metadata: dict[str, object] | None = None,
    ) -> tuple[str, bytes]:
        explicit_lead_ids = lead_ids is not None
        if lead_ids is None:
            leads = self.lead_repository.list_all_leads(filters)
        else:
            requested_ids = list(dict.fromkeys(int(lead_id) for lead_id in lead_ids))
            leads = self.lead_repository.list_export_leads_by_ids(requested_ids, blocked=filters.blocked)
        scope_metadata = dict(scope_metadata or {})
        if explicit_lead_ids:
            scope_metadata.setdefault("scope_type", "explicit_lead_ids")

        leads_df = pd.DataFrame([self._lead_row(lead) for lead in leads])
        metadata_df = pd.DataFrame(
            self._metadata_rows(
                leads,
                leads_df,
                filters,
                scope_label=scope_label,
                scope_metadata=scope_metadata,
            )
        )

        output = BytesIO()
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            self._write_sheet(writer, "Leads", leads_df)
            self._write_sheet(writer, "Metadata", metadata_df)

            workbook = writer.book
            self._format_sheet(workbook["Leads"], highlight_missing_contacts=True)
            self._format_sheet(workbook["Metadata"], freeze_headers=False)

        filename = f"lead_export_{datetime.now(timezone.utc).strftime('%Y%m%d_%H%M%S')}.xlsx"
        return filename, output.getvalue()

    def _write_sheet(self, writer: pd.ExcelWriter, sheet_name: str, dataframe: pd.DataFrame) -> None:
        frame = dataframe.copy()
        expected_columns = SHEET_COLUMNS.get(sheet_name)
        if expected_columns:
            frame = frame.reindex(columns=expected_columns)
        if frame.empty and expected_columns:
            frame = pd.DataFrame(columns=expected_columns)
        frame.to_excel(writer, sheet_name=sheet_name, index=False)

    def _format_sheet(self, worksheet, *, highlight_missing_contacts: bool = False, freeze_headers: bool = True) -> None:
        max_row = worksheet.max_row
        max_col = worksheet.max_column
        if max_row >= 1:
            for cell in worksheet[1]:
                cell.fill = HEADER_FILL
                cell.font = HEADER_FONT
            if freeze_headers:
                worksheet.freeze_panes = "A2"
            worksheet.auto_filter.ref = worksheet.dimensions

        for column_cells in worksheet.columns:
            values = [str(cell.value) if cell.value is not None else "" for cell in column_cells]
            max_length = max((len(value) for value in values), default=0)
            worksheet.column_dimensions[column_cells[0].column_letter].width = min(max_length + 2, 45)

        if highlight_missing_contacts and max_row > 1:
            headers = {worksheet.cell(row=1, column=index).value: index for index in range(1, max_col + 1)}
            for header_name in ["email", "phone", "whatsapp"]:
                column_index = headers.get(header_name)
                if not column_index:
                    continue
                column_letter = get_column_letter(column_index)
                worksheet.conditional_formatting.add(
                    f"{column_letter}2:{column_letter}{max_row}",
                    FormulaRule(
                        formula=[f'LEN(TRIM({column_letter}2))=0'],
                        fill=MISSING_FILL,
                    ),
                )

    def _lead_row(self, lead) -> dict[str, object]:
        email = self._best_export_email(lead)
        phone = self._best_contact_value(lead, ContactType.PHONE, lead.phone)
        whatsapp = self._best_contact_value(lead, ContactType.WHATSAPP, lead.whatsapp)
        instagram = self._best_contact_value(lead, ContactType.INSTAGRAM, lead.instagram)
        street_name, street_number = split_street_and_number(
            lead.address,
            neighborhood=lead.neighborhood,
            city=lead.city,
            state=lead.state,
            postal_code=lead.postal_code,
        )
        return {
            "business_name": lead.business_name,
            "category": lead.category,
            "city": lead.city,
            "state": normalize_brazilian_state(lead.state),
            "address": lead.address,
            "street": street_name,
            "number": street_number,
            "postal_code": lead.postal_code,
            "phone": phone,
            "whatsapp": whatsapp,
            "email": email,
            "instagram": instagram,
            "website": self._client_facing_website(lead.website),
            "google_maps_url": lead.google_maps_url,
            "source": lead.lead_source_type.value if lead.lead_source_type else None,
            "status": lead.status.value if lead.status else None,
            "notes": lead.notes,
            "neighborhood": lead.neighborhood,
            "assigned_owner": lead.assigned_sales_rep.name if lead.assigned_sales_rep else None,
            "market_segment": lead.market_segment.name if lead.market_segment else None,
            "market_subsegment": lead.market_subsegment.name if lead.market_subsegment else None,
            "lead_score": lead.lead_score,
            "rating": None,
            "review_count": None,
        }

    @classmethod
    def _best_contact_value(cls, lead, contact_type: ContactType, fallback: str | None) -> str | None:
        contacts = cls._ranked_contacts(lead, contact_type)
        if not contacts:
            return fallback
        best = contacts[0]
        return cls._contact_value(best) or fallback

    @classmethod
    def _best_export_email(cls, lead) -> str | None:
        for contact in cls._ranked_contacts(lead, ContactType.EMAIL):
            value = cls._contact_value(contact)
            if is_probable_client_facing_email(value):
                return value
        if is_probable_client_facing_email(lead.email):
            return lead.email
        return None

    @classmethod
    def _ranked_contacts(cls, lead, contact_type: ContactType) -> list:
        contacts = [
            contact
            for contact in lead.contacts
            if cls._contact_type_matches(contact.contact_type, contact_type) and cls._contact_value(contact)
        ]
        return sorted(
            contacts,
            key=lambda contact: (
                1 if contact.is_primary else 0,
                float(contact.confidence or 0),
                contact.updated_at.timestamp() if contact.updated_at else 0,
                contact.id or 0,
            ),
            reverse=True,
        )

    @staticmethod
    def _contact_value(contact) -> str | None:
        return getattr(contact, "normalized_value", None) or getattr(contact, "raw_value", None)

    @classmethod
    def _contact_type_matches(cls, actual: object, expected: ContactType) -> bool:
        actual_token = cls._contact_type_token(actual)
        expected_tokens = {
            cls._contact_type_token(expected),
            cls._contact_type_token(expected.name),
            cls._contact_type_token(expected.value),
        }
        return bool(actual_token and actual_token in expected_tokens)

    @staticmethod
    def _contact_type_token(value: object) -> str:
        raw_value = getattr(value, "value", value)
        text = str(raw_value or "").strip()
        if "." in text:
            text = text.rsplit(".", 1)[-1]
        return text.lower()

    @staticmethod
    def _client_facing_website(value: str | None) -> str | None:
        if not is_probable_business_website(value):
            return None
        return canonicalize_url(value)

    @staticmethod
    def _metadata_rows(
        leads: list,
        leads_df: pd.DataFrame,
        filters: LeadListFilters,
        *,
        scope_label: str | None = None,
        scope_metadata: dict[str, object] | None = None,
    ) -> list[dict[str, object]]:
        scope = dict(scope_metadata or {})
        scope_type = str(scope.get("scope_type") or "all_leads_matching_advanced_filters")
        scope_label = scope_label or str(scope.get("scope_label") or "All leads matching advanced filters")
        scope["scope_type"] = scope_type
        scope["scope_label"] = scope_label
        scope["lead_count"] = int(len(leads))

        duplicate_count = sum(1 for lead in leads if lead.is_duplicate)
        total_with_email = int(leads_df["email"].fillna("").astype(str).str.len().gt(0).sum()) if not leads_df.empty else 0
        total_with_whatsapp = (
            int(leads_df["whatsapp"].fillna("").astype(str).str.len().gt(0).sum()) if not leads_df.empty else 0
        )
        total_with_website = (
            int(leads_df["website"].fillna("").astype(str).str.len().gt(0).sum()) if not leads_df.empty else 0
        )
        score_min = int(leads_df["lead_score"].min()) if not leads_df.empty else 0
        score_avg = float(leads_df["lead_score"].mean()) if not leads_df.empty else 0.0
        score_max = int(leads_df["lead_score"].max()) if not leads_df.empty else 0

        return [
            {"metric": "export_timestamp_utc", "value": datetime.now(timezone.utc).isoformat()},
            {"metric": "scope_type", "value": scope_type},
            {"metric": "scope_label", "value": scope_label},
            {"metric": "lead_count", "value": int(len(leads))},
            {"metric": "scope_metadata", "value": json.dumps(scope, ensure_ascii=True, default=str)},
            {"metric": "applied_filters", "value": json.dumps(filters.model_dump(mode="json"), ensure_ascii=True)},
            {"metric": "total_lead_count", "value": int(len(leads))},
            {"metric": "total_duplicate_count", "value": duplicate_count},
            {"metric": "total_with_email", "value": total_with_email},
            {"metric": "total_with_whatsapp", "value": total_with_whatsapp},
            {"metric": "total_with_website", "value": total_with_website},
            {"metric": "score_min", "value": score_min},
            {"metric": "score_avg", "value": round(score_avg, 2)},
            {"metric": "score_max", "value": score_max},
        ]
